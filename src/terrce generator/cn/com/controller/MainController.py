'''
Created on Aug 30, 2015

@author: zhihan
'''
import os
from cn.com.assistant import DataMapping
from cn.com.controller import VersionController
from openpyxl.conftest import Workbook
from openpyxl.reader.excel import load_workbook

wbDir = DataMapping.OUTPUT_TOP+DataMapping.PATH_SEPERATOR+DataMapping.OUTPUT_FILE_NAME

def main():
    products = readProductName()
    processProducts(products)


'''
@note: output head of tables
'''
def outputHead(ws):
    ws["A1:B1"] = "version"
    
    return 0
    
    
'''
@note: read all product names in specific directory
'''
def readProductName():
    return os.listdir(DataMapping.TOP_DIR)

'''
@author: zhihan
@param products: all product name
'''
def processProducts(products):
    wb = createWb()
#     for product in products:
#         productDir = DataMapping.TOP_DIR + DataMapping.PATH_SEPERATOR+product
#         ws = createWs(product,wb)
#         versions = os.listdir(productDir)
#         processVersions(product,versions,ws)
    wb.save(wbDir)
    
'''
@attention: to process every version of product
@param product: name of a specific product
@param versions: all versions of product 
@param ws: responding worksheet of product 
'''
def processVersions(product,versions,ws):
    for version in versions:
        VersionController.VersionController(product,version,ws)

'''
@note: create sheet in workbook and sheet name is sheetName
@param sheetName: the sheet name   
'''       
def createWs(sheetName,wb):
    ws = wb.create_sheet()
    ws.title = sheetName
    return ws
    
'''
@note: create workbook named in DataMapping
'''
def createWb():
    if(not os.path.exists(DataMapping.OUTPUT_TOP)):
        os.makedirs(DataMapping.OUTPUT_TOP)
    if(not os.path.isfile(wbDir)):
        wb = Workbook()
    else:
        wb = load_workbook(wbDir)
    return wb

    
    
    
    