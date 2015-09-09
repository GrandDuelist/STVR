'''
Created on Sep 7, 2015

@author: zhihan
'''
import os;
import re
import string
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from Main import DataMapping
def readFromFile(fileName):
    fp = open(fileName,"r+")
    result = []
    for line in fp:
       # search_pat = re.compile(r'(\s+)((\d+)(.?)(\d*))(\s*)')
        search_pat = re.compile(r'\s+(\d+.\d*)\s+')
        start = string.find(line,"]")+2
        end = string.find(line,"\n")
      
        line = line[start:end]
        currentResult = line.split(r" ")
        if DataMapping.MEDIUM not in currentResult:
            result.extend(currentResult)
        else:
            return result
    return result


def writeToExcel(fileName,sheetName,colChar,contents,colName):
    if(not os.path.isfile(fileName)):
        wb = Workbook()
    
    else:
        wb = load_workbook(fileName)
    sheetNames = wb.get_sheet_names()
    
    if sheetName in sheetNames:
        ws = wb.get_sheet_by_name(sheetName)
    else:
        ws = wb.create_sheet()
        ws.title = sheetName
        
    ws[(colChar+str(1))] = colName
    i = 2
    for content in contents:
        cellIndex = colChar + str(i)
        ws[cellIndex] = content
        i = i+1
    wb.save(fileName)
    
def tranOneFile(fileName,colChar,colName,outputFile):
    baseFileName = os.path.basename(fileName)
    baseFileName = os.path.splitext(baseFileName)[0]
    contents = readFromFile(fileName)
    writeToExcel(outputFile,baseFileName,colChar, contents, colName)
    
def tranOneProduct(topDir,outputFile):
    methods = os.listdir(topDir)
    initColChar = 'A'
    i = 0 
    for method in methods:
        currentColChar = charAdd(initColChar,i)
        methodDir = os.path.join(topDir,method)
        versions = os.listdir(methodDir)
        for version in versions:
            fileDir = os.path.join(methodDir,version)
            print fileDir
            tranOneFile(fileDir,currentColChar,method,outputFile)
        i=i+1
        
def charAdd(char,number):
    return chr(ord(char)+number)